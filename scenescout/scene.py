"""DelawareScene.com reference scrapers.

scene_orgs:     the full organization directory (one page) -> name -> VenueID map.
scene_listings: what the calendar currently lists, from two origins:
  - 'xlsx': assets/DelawareScene Currently Listed Events.xlsx (bootstrap snapshot)
  - 'live': day pages at /search/?start=YYYY-MM-DD (the only server-side
    enumeration that works; the form's "Anytime!" option is JS-only, verified
    2026-08-18).

Day-page markup (verified 2026-08-18):
  <a href="https://delawarescene.com/event/{id}/{slug}" class="cell event">
    <h3>Title</h3>
    <div class="info"><ul class="items"><li>Venue</li><li>Thursday, August 20</li></ul></div>
  </a>
Ongoing events repeat on every day page with <li>Through October 4</li>.
"""

from __future__ import annotations

import re
from datetime import date, timedelta
from pathlib import Path

from bs4 import BeautifulSoup
from dateutil import parser as dateparser

from . import db, http

BASE = "https://delawarescene.com"
DIRECTORY_URL = f"{BASE}/about/directory.php"
ORG_RE = re.compile(r"/organization/(\d+)/([\w-]+)")
EVENT_RE = re.compile(r"/event/(\d+)/[\w-]*")
_DATE_ITEM_RE = re.compile(
    r"^(Mon|Tues|Wednes|Thurs|Fri|Satur|Sun)day,\s", re.IGNORECASE
)

ASSETS = Path(__file__).resolve().parent.parent / "assets"
LISTED_XLSX = ASSETS / "DelawareScene Currently Listed Events.xlsx"


def scrape_orgs(conn) -> int:
    """Fetch the directory page and upsert every organization. Returns count."""
    status, content, _ = http.fetch(DIRECTORY_URL, conn=conn, max_age_hours=24)
    if status != 200:
        raise RuntimeError(f"directory fetch failed: HTTP {status}")
    soup = BeautifulSoup(content, "html.parser")
    seen = {}
    for a in soup.find_all("a", href=ORG_RE):
        m = ORG_RE.search(a["href"])
        name = a.get_text(" ", strip=True)
        if not name:
            continue
        scene_id = int(m.group(1))
        # Directory repeats links (nav letters etc.); first name wins.
        seen.setdefault(scene_id, (name, m.group(2)))
    for scene_id, (name, slug) in seen.items():
        conn.execute(
            "INSERT OR REPLACE INTO scene_orgs (scene_id, name, slug, url) "
            "VALUES (?, ?, ?, ?)",
            (scene_id, name, slug, f"{BASE}/organization/{scene_id}/{slug}"),
        )
    conn.commit()
    return len(seen)


def bootstrap_from_xlsx(conn, path: Path = LISTED_XLSX) -> int:
    """Load the DDOA 'Currently Listed Events' export as baseline listings."""
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    assert [str(h).strip().lower() for h in header[:4]] == [
        "title", "venue", "start", "end",
    ], f"unexpected header in {path.name}: {header!r}"
    n = 0
    for title, venue, start, end in (r[:4] for r in rows):
        if not title:
            continue
        # venue is COALESCE'd to '' because SQLite treats NULLs as distinct in
        # a unique index, so NULL-venue rows would duplicate on every run.
        conn.execute(
            "INSERT OR IGNORE INTO scene_listings "
            "(scene_event_id, title, venue, start_date, end_date, is_range, url, origin, scraped_at) "
            "VALUES (NULL, ?, ?, ?, ?, 1, NULL, 'xlsx', ?)",
            (
                str(title).strip(),
                (str(venue).strip() if venue else ""),
                _iso(start),
                _iso(end),
                db.now(),
            ),
        )
        n += 1
    conn.commit()
    return n


def scrape_listings(conn, start: date = None, days: int = 120) -> int:
    """Walk /search/?start= day pages; upsert by scene event id. Returns count."""
    start = start or date.today()
    found = {}
    for i in range(days):
        day = start + timedelta(days=i)
        url = f"{BASE}/search/?start={day.isoformat()}"
        status, content, _ = http.fetch(url, conn=conn, max_age_hours=24)
        if status != 200:
            continue
        for ev in _parse_day_page(content, day):
            prior = found.get(ev["id"])
            if prior is None:
                ev["dates"] = [ev["start_date"]]
                # A "Through <date>" marker means a genuine continuous run;
                # without one, each sighting is a discrete occurrence.
                ev["is_range"] = 1 if ev["end_date"] else 0
                found[ev["id"]] = ev
                continue
            prior["dates"].append(ev["start_date"])
            if ev["end_date"]:
                prior["is_range"] = 1
                if not prior["end_date"] or ev["end_date"] > prior["end_date"]:
                    prior["end_date"] = ev["end_date"]
    for ev in found.values():
        dates = sorted(set(ev["dates"]))
        conn.execute(
            "INSERT INTO scene_listings "
            "(scene_event_id, title, venue, start_date, end_date, dates, is_range, "
            " url, origin, scraped_at) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'live', ?) "
            "ON CONFLICT(scene_event_id) WHERE scene_event_id IS NOT NULL DO UPDATE SET "
            "title=excluded.title, venue=excluded.venue, "
            "start_date=excluded.start_date, end_date=excluded.end_date, "
            "dates=excluded.dates, is_range=excluded.is_range, "
            "scraped_at=excluded.scraped_at",
            (
                ev["id"], ev["title"], ev["venue"],
                dates[0], ev["end_date"] or (dates[-1] if len(dates) > 1 else None),
                db.to_json(dates), ev["is_range"], ev["url"], db.now(),
            ),
        )
    conn.commit()
    return len(found)


def _parse_day_page(content: bytes, day: date):
    soup = BeautifulSoup(content, "html.parser")
    for a in soup.select("a.cell.event"):
        m = EVENT_RE.search(a.get("href", ""))
        if not m:
            continue
        h3 = a.find("h3")
        items = [li.get_text(" ", strip=True) for li in a.select("ul.items li")]
        # Items are venue and/or a date line ("Tuesday, August 18") and/or
        # "Through October 4"; venue-less events exist, so classify each item.
        venue = None
        end_date = None
        for it in items:
            low = it.lower()
            if low.startswith("through"):
                end_date = _parse_through(it, day)
            elif _DATE_ITEM_RE.match(it):
                continue
            elif venue is None:
                venue = it
        yield {
            "id": int(m.group(1)),
            "title": h3.get_text(" ", strip=True) if h3 else None,
            "venue": venue,
            "start_date": day.isoformat(),
            "end_date": end_date,
            "url": a["href"],
        }


def _parse_through(text: str, day: date):
    # "Through October 4" (year implied, may wrap) or "Through January 3, 2027"
    raw = text[len("through"):].strip(" :")
    try:
        d = dateparser.parse(raw, default=dateparser.parse(f"{day.year}-01-01")).date()
        if d < day and not re.search(r"\d{4}", raw):
            d = d.replace(year=day.year + 1)
        return d.isoformat()
    except (ValueError, OverflowError):
        return None


def _iso(value):
    if value is None or value == "":
        return None
    if hasattr(value, "date"):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    try:
        return dateparser.parse(str(value)).date().isoformat()
    except (ValueError, OverflowError):
        return None
