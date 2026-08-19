"""Mock DelawareScene bulk-upload portal.

A small Flask app that mimics the DDOA intake workflow end to end:
upload the 13-column bulk-upload .xlsx, run the same validators the exporter
uses, and render accepted events as a DelawareScene-style calendar list.

    .venv/bin/python mock_site/app.py     # http://127.0.0.1:5050
"""

from __future__ import annotations

import os
import re
import sqlite3
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from flask import Flask, redirect, render_template_string, request, url_for  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

from scenescout.dedupe import DUPE_SCORE, TITLE_ONLY_DUPE, score_pair  # noqa: E402
from scenescout.export import HEADERS, validate_row  # noqa: E402
from scenescout.normalize import norm_title  # noqa: E402

DB = ROOT / "data" / "mock_site.db"
app = Flask(__name__)

CATEGORY_NAMES = {
    1: "Attractions", 2: "Dance", 3: "Festivals & Special Events", 4: "Film",
    5: "Free", 6: "Kids & Family Friendly", 7: "Lectures & Workshops",
    8: "Literature & Poetry", 9: "Music", 10: "Theater & Performance",
    11: "Visual Arts", 12: "Bands", 13: "Choral", 14: "Classical/Opera",
    15: "Country/Folk/Bluegrass", 16: "Hip Hop/R&B", 17: "Jazz/Blues",
    18: "Rock/Pop", 19: "World", 20: "Comedy/Drama", 21: "Musical",
    22: "Variety", 23: "Art Centers", 24: "Art, Antiques & Craft Shows",
    25: "Art Tours", 26: "Exhibitions", 27: "Galleries", 28: "Museums",
    29: "Public Art",
}

# Events are coloured by discipline rather than by individual category, so a
# jazz gig and a choral concert read as the same kind of thing at a glance.
# Subcategories inherit their parent's colour.
FAMILIES = [
    ("music", "Music", "#3D5A99", {9, 12, 13, 14, 15, 16, 17, 18, 19}),
    ("theater", "Theater & Performance", "#7A4E8C", {10, 20, 21, 22}),
    ("visual", "Visual Arts", "#B5761F", {11, 23, 24, 25, 26, 27, 28, 29}),
    ("dance", "Dance", "#B04A6E", {2}),
    ("film", "Film", "#2E7D74", {4}),
    ("lit", "Literature & Poetry", "#4A7C4E", {8}),
    ("festival", "Festivals", "#C0552F", {3}),
    ("talks", "Lectures & Workshops", "#5B6673", {7}),
    ("kids", "Kids & Family", "#2C7A94", {6}),
    ("attractions", "Attractions", "#8A6A45", {1}),
]
FAMILY_COLORS = {key: color for key, _, color, _ in FAMILIES}
OTHER_COLOR = "#6B7280"


def family_of(category_ids):
    """First matching discipline; 'Free' (5) is a price flag, not a discipline."""
    for cid in category_ids:
        for key, _, _, members in FAMILIES:
            if cid in members:
                return key
    return "other"

CALENDAR_PAGE = """<!doctype html><meta charset="utf-8">
<title>MockScene — calendar</title>
<style>
  :root {
    --blue:#2F5384; --buff:#C9A85C; --bg:#F7F5EE; --surface:#FDFCF7;
    --ink:#212C39; --muted:#5B6673; --line:#E2DCCB;
  }
  * { box-sizing: border-box; }
  body { font-family: system-ui, -apple-system, sans-serif; background: var(--bg);
         color: var(--ink); margin: 0; line-height: 1.45; }
  a { color: inherit; }

  header { background: var(--blue); color: #fff; padding: 0.85rem 1.5rem;
           display: flex; align-items: center; gap: 1.5rem; flex-wrap: wrap; }
  header h1 { margin: 0; font-size: 1.15rem; font-weight: 600; }
  header h1 span { color: var(--buff); }
  header nav { margin-left: auto; display: flex; gap: 1rem; font-size: 0.9rem; }
  header nav a { color: #cfe0f5; text-decoration: none; padding-bottom: 2px; }
  header nav a.on { color: #fff; border-bottom: 2px solid var(--buff); }

  .toolbar { display: flex; align-items: center; gap: 0.75rem; flex-wrap: wrap;
             padding: 0.9rem 1.5rem; border-bottom: 1px solid var(--line);
             background: var(--surface); position: sticky; top: 0; z-index: 20; }
  .period { font-size: 1.15rem; font-weight: 600; min-width: 15rem; }
  .btn { background: var(--surface); border: 1px solid var(--line); color: var(--ink);
         border-radius: 6px; padding: 0.35rem 0.7rem; cursor: pointer; font: inherit;
         font-size: 0.85rem; }
  .btn:hover { background: #F1ECDD; }
  .btn:focus-visible { outline: 2px solid var(--blue); outline-offset: 1px; }
  .seg { display: flex; border: 1px solid var(--line); border-radius: 6px; overflow: hidden; }
  .seg button { border: 0; background: var(--surface); padding: 0.35rem 0.85rem;
                cursor: pointer; font: inherit; font-size: 0.85rem; }
  .seg button.on { background: var(--blue); color: #fff; }
  .count { margin-left: auto; font-size: 0.85rem; color: var(--muted); }
  .search { border: 1px solid var(--line); border-radius: 6px; padding: 0.35rem 0.6rem;
            font: inherit; font-size: 0.85rem; background: var(--surface); width: 13rem; }

  .legend { display: flex; gap: 0.4rem; flex-wrap: wrap; padding: 0.65rem 1.5rem;
            border-bottom: 1px solid var(--line); background: var(--bg);
            align-items: center; }
  .legend .lab { font-size: 0.72rem; text-transform: uppercase; letter-spacing: 0.06em;
                 color: var(--muted); margin-right: 0.2rem; }
  .legend .sep { width: 1px; align-self: stretch; background: var(--line); margin: 0 0.4rem; }
  .legend .all { font-size: 0.72rem; color: var(--blue); background: none; border: 0;
                 cursor: pointer; font-family: inherit; text-decoration: underline;
                 text-underline-offset: 2px; padding: 0.2rem 0.15rem; }

  /* Showing: filled with the discipline's own colour, so the legend doubles as
     the colour key. Hidden: hollow and desaturated — a different shape of chip,
     not just a fainter one. */
  .chip { display: inline-flex; align-items: center; gap: 0.4rem; font-size: 0.75rem;
          border: 1px solid; border-radius: 99px; padding: 0.22rem 0.7rem;
          cursor: pointer; user-select: none; font-weight: 600;
          background: color-mix(in srgb, var(--c) 14%, #fff);
          border-color: color-mix(in srgb, var(--c) 42%, #fff);
          /* 65% keeps each chip recognisably its own hue while staying above
             the 4.5:1 contrast floor — amber is light enough that a higher
             mix falls under it. */
          color: color-mix(in srgb, var(--c) 65%, #16202C);
          transition: background .12s ease, border-color .12s ease, color .12s ease; }
  .chip .dot { width: 9px; height: 9px; border-radius: 50%; background: var(--c);
               box-shadow: 0 0 0 2px color-mix(in srgb, var(--c) 20%, #fff); }
  .chip .n { font-weight: 500; opacity: 0.7; font-variant-numeric: tabular-nums; }
  .chip:hover { background: color-mix(in srgb, var(--c) 22%, #fff); }
  .chip:focus-visible { outline: 2px solid var(--blue); outline-offset: 2px; }

  .chip.off { background: transparent; border-color: #D8D2C0; color: #949AA1;
              font-weight: 500; }
  .chip.off .dot { background: transparent; box-shadow: none;
                   border: 1.5px solid #C3BCA9; width: 8px; height: 8px; }
  .chip.off:hover { background: #F1ECDF; border-color: #C9C2AF; color: var(--ink); }

  main { padding: 1.1rem 1.5rem 4rem; }

  /* ---- month ---- */
  /* minmax(0, 1fr) rather than 1fr: a bare 1fr is minmax(auto, 1fr), so the
     nowrap event titles would set a huge minimum width on every column, the
     grid would overflow, and overflow:hidden would clip most of the week. */
  /* 1px side padding mirrors the month grid's border so the weekday labels
     line up with the columns beneath them. */
  .dow { display: grid; grid-template-columns: repeat(7, minmax(0, 1fr)); gap: 1px;
         font-size: 0.72rem; text-transform: uppercase; letter-spacing: 0.06em;
         color: var(--muted); margin-bottom: 4px; padding: 0 1px; }
  .dow div { padding: 0 0.4rem; }
  .month { display: grid; grid-template-columns: repeat(7, minmax(0, 1fr)); gap: 1px;
           background: var(--line); border: 1px solid var(--line); border-radius: 8px;
           overflow: hidden; }
  .cell { background: var(--surface); min-height: 7.5rem; padding: 0.3rem 0.35rem;
          display: flex; flex-direction: column; gap: 2px;
          min-width: 0; overflow: hidden; }
  .cell.dim { background: #F4F1E6; }
  .cell .dnum { font-size: 0.78rem; color: var(--muted); font-variant-numeric: tabular-nums; }
  .cell.today .dnum { background: var(--blue); color: #fff; border-radius: 50%;
                      width: 1.4rem; height: 1.4rem; display: grid; place-items: center;
                      font-weight: 600; }
  .ev { font-size: 0.73rem; line-height: 1.25; padding: 2px 5px; border-radius: 4px;
        border-left: 3px solid var(--c); background: color-mix(in srgb, var(--c) 12%, white);
        color: #1c2530; cursor: pointer; overflow: hidden; text-overflow: ellipsis;
        white-space: nowrap; min-width: 0; max-width: 100%; }
  .ev:hover { background: color-mix(in srgb, var(--c) 22%, white); }
  .ev .t { color: var(--muted); font-variant-numeric: tabular-nums; }
  /* A left notch marks events SceneScout contributed, so the calendar shows
     at a glance what the pipeline added to what was already there. */
  .ev.new { border-left-style: double; border-left-width: 5px; }
  .more { font-size: 0.7rem; color: var(--muted); cursor: pointer; padding-left: 5px; }

  /* ---- week / day ---- */
  .grid { display: grid; background: var(--line); border: 1px solid var(--line);
          border-radius: 8px; overflow: hidden; gap: 1px; }
  .grid .head { background: var(--surface); padding: 0.4rem; text-align: center;
                font-size: 0.8rem; }
  .grid .head b { display: block; font-size: 1.1rem; font-variant-numeric: tabular-nums; }
  .grid .head.today b { color: var(--blue); }
  .daycol { background: var(--surface); display: flex; flex-direction: column;
            min-width: 0; overflow: hidden; }
  .col { background: var(--surface); position: relative; min-height: 62.4rem; flex: 1; }
  .hour { position: absolute; left: 0; right: 0; border-top: 1px solid #EFEADB; }
  .hourlab { position: absolute; left: 0; width: 100%; font-size: 0.68rem;
             color: var(--muted); padding-left: 4px; }
  .tev { position: absolute; left: 3px; right: 3px; border-radius: 4px; padding: 2px 5px;
         font-size: 0.72rem; line-height: 1.2; overflow: hidden; cursor: pointer;
         border-left: 3px solid var(--c); background: color-mix(in srgb, var(--c) 14%, white); }
  .allday { padding: 3px; display: flex; flex-direction: column; gap: 2px;
            min-height: 2.1rem; border-bottom: 1px solid var(--line); }

  /* ---- detail ---- */
  dialog { border: 1px solid var(--line); border-radius: 10px; padding: 0; max-width: 34rem;
           background: var(--surface); color: var(--ink); }
  dialog::backdrop { background: rgba(20,28,40,0.45); }
  .d-head { padding: 1rem 1.2rem 0.7rem; border-bottom: 1px solid var(--line);
            border-top: 4px solid var(--c); }
  .d-head h3 { margin: 0 0 0.2rem; font-size: 1.1rem; }
  .d-body { padding: 0.9rem 1.2rem 1.2rem; font-size: 0.9rem; }
  .d-body dt { font-size: 0.72rem; text-transform: uppercase; letter-spacing: 0.05em;
               color: var(--muted); margin-top: 0.7rem; }
  .d-body dd { margin: 0.1rem 0 0; }
  .tag { display: inline-block; font-size: 0.7rem; border-radius: 99px; padding: 0.1rem 0.5rem;
         background: #EEF2F8; color: var(--blue); margin: 0.15rem 0.15rem 0 0; }
  .empty { text-align: center; color: var(--muted); padding: 4rem 1rem; }
</style>

<header>
  <h1>Mock<span>Scene</span></h1>
  <nav>
    <a href="{{ url_for('index') }}">Bulk upload</a>
    <a href="{{ url_for('calendar_view') }}" class="on">Calendar</a>
  </nav>
</header>

<div class="toolbar">
  <button class="btn" id="today">Today</button>
  <button class="btn" id="prev" aria-label="Previous">&#8249;</button>
  <button class="btn" id="next" aria-label="Next">&#8250;</button>
  <span class="period" id="period"></span>
  <div class="seg" id="views">
    <button data-v="month" class="on">Month</button>
    <button data-v="week">Week</button>
    <button data-v="day">Day</button>
  </div>
  <div class="seg" id="origins">
    <button data-o="all" class="on">All</button>
    <button data-o="listed">Already listed</button>
    <button data-o="scenescout">Found by SceneScout</button>
  </div>
  <input class="search" id="q" type="search" placeholder="Filter by title or venue">
  <span class="count" id="count"></span>
</div>

<div class="legend" id="legend"></div>
<main id="view"></main>

<dialog id="detail"><div class="d-head" id="dhead"></div><div class="d-body" id="dbody"></div>
  <div style="padding:0 1.2rem 1.1rem"><button class="btn" id="dclose">Close</button></div>
</dialog>

<script>
const EVENTS = {{ events_json | safe }};
const FAMILIES = {{ families_json | safe }};
const OTHER = "{{ other_color }}";
const colorOf = f => FAMILIES[f] ? FAMILIES[f].color : OTHER;

let view = "month";
// Open on today; fall back to the first event if the calendar is all future.
let cursor = new Date();
const hidden = new Set();
let query = "";

const pad = n => String(n).padStart(2, "0");
const iso = d => `${d.getFullYear()}-${pad(d.getMonth()+1)}-${pad(d.getDate())}`;
const parse = s => new Date(s + "T12:00:00");
const addDays = (d, n) => { const x = new Date(d); x.setDate(x.getDate()+n); return x; };
const startOfWeek = d => addDays(d, -d.getDay());
const sameDay = (a, b) => iso(a) === iso(b);
const TODAY = new Date();

let originFilter = "all";
function visible(e) {
  if (hidden.has(e.family)) return false;
  if (originFilter !== "all" && e.origin !== originFilter) return false;
  if (!query) return true;
  const q = query.toLowerCase();
  return (e.title || "").toLowerCase().includes(q) || (e.venue || "").toLowerCase().includes(q);
}
// An exhibit occupies every day between its start and end.
function onDay(e, dstr) { return e.start <= dstr && (e.end || e.start) >= dstr; }
function eventsOn(d) {
  const s = iso(d);
  return EVENTS.filter(e => visible(e) && onDay(e, s))
               .sort((a, b) => (a.minutes ?? 1e9) - (b.minutes ?? 1e9));
}

function renderLegend() {
  const el = document.getElementById("legend");
  const otherCount = EVENTS.filter(e => e.family === "other").length;
  const entries = Object.entries(FAMILIES);
  if (otherCount) entries.push(["other", {label: "Uncategorised", color: OTHER, count: otherCount}]);

  const chip = ([k, f]) => {
    const on = !hidden.has(k);
    return `<span class="chip ${on ? "" : "off"}" data-f="${k}" style="--c:${f.color}"
                  role="checkbox" tabindex="0" aria-checked="${on}"
                  title="${on ? "Showing" : "Hidden"} — click to toggle">
              <span class="dot"></span>${f.label} <span class="n">${f.count}</span></span>`;
  };
  el.innerHTML = `<span class="lab">Discipline</span>` + entries.map(chip).join("")
    + `<span class="sep"></span>`
    + `<button class="all" data-act="show">Show all</button>`
    + `<button class="all" data-act="hide">Hide all</button>`;

  const toggle = c => {
    const f = c.dataset.f;
    hidden.has(f) ? hidden.delete(f) : hidden.add(f);
    renderLegend();
    render();
  };
  el.querySelectorAll(".chip").forEach(c => {
    c.onclick = () => toggle(c);
    c.onkeydown = ev => {
      if (ev.key === "Enter" || ev.key === " ") { ev.preventDefault(); toggle(c); }
    };
  });
  el.querySelectorAll(".all").forEach(b => b.onclick = () => {
    hidden.clear();
    if (b.dataset.act === "hide") entries.forEach(([k]) => hidden.add(k));
    renderLegend();
    render();
  });
}

function chip(e) {
  const t = e.timeLabel ? `<span class="t">${e.timeLabel}</span> ` : "";
  const isNew = e.origin !== "listed" ? " new" : "";
  return `<div class="ev${isNew}" style="--c:${colorOf(e.family)}" data-id="${e.id}"
            title="${esc(e.title)}">${t}${esc(e.title)}</div>`;
}
const esc = s => String(s ?? "").replace(/[&<>"]/g, c =>
  ({"&":"&amp;","<":"&lt;",">":"&gt;",'"':"&quot;"}[c]));

function renderMonth() {
  const first = new Date(cursor.getFullYear(), cursor.getMonth(), 1);
  const start = startOfWeek(first);
  let html = `<div class="dow">${["Sun","Mon","Tue","Wed","Thu","Fri","Sat"]
    .map(d => `<div>${d}</div>`).join("")}</div><div class="month">`;
  for (let i = 0; i < 42; i++) {
    const d = addDays(start, i);
    const evs = eventsOn(d);
    const dim = d.getMonth() !== cursor.getMonth() ? " dim" : "";
    const today = sameDay(d, TODAY) ? " today" : "";
    const shown = evs.slice(0, 3).map(chip).join("");
    const more = evs.length > 3
      ? `<div class="more" data-day="${iso(d)}">+${evs.length - 3} more</div>` : "";
    html += `<div class="cell${dim}${today}"><div class="dnum">${d.getDate()}</div>${shown}${more}</div>`;
  }
  return html + "</div>";
}

// One grid cell per day: an all-day strip on top, then the timed area.
// Returning two siblings here would make each day occupy two columns.
function dayCell(d) {
  const evs = eventsOn(d);
  const allday = evs.filter(e => e.minutes === null);
  const timed = evs.filter(e => e.minutes !== null);
  let slots = "";
  for (let h = 0; h < 24; h++) slots += `<div class="hour" style="top:${h * 2.6}rem"></div>`;
  // Events at the same time share the width instead of hiding each other.
  timed.forEach((e, i) => {
    const clash = timed.filter(x => Math.abs(x.minutes - e.minutes) < 45);
    const idx = clash.indexOf(e), n = Math.max(clash.length, 1);
    const w = 100 / n;
    slots += `<div class="tev" data-id="${e.id}"
        style="--c:${colorOf(e.family)};top:${(e.minutes / 60) * 2.6}rem;height:2.4rem;
               left:calc(${idx * w}% + 3px);width:calc(${w}% - 6px)">
        <b>${e.timeLabel}</b> ${esc(e.title)}</div>`;
  });
  return `<div class="daycol">
      <div class="allday">${allday.map(chip).join("")}</div>
      <div class="col">${slots}</div>
    </div>`;
}

function renderDays(days) {
  const dow = ["Sun","Mon","Tue","Wed","Thu","Fri","Sat"];
  const heads = days.map(d =>
    `<div class="head${sameDay(d, TODAY) ? " today" : ""}">${dow[d.getDay()]}
       <b>${d.getDate()}</b></div>`).join("");
  const hours = Array.from({length: 24}, (_, h) =>
    `<div class="hourlab" style="top:${h * 2.6 - 0.4}rem">${h % 12 || 12}${h < 12 ? "a" : "p"}</div>`
  ).join("");
  // minmax(0,1fr) for the same reason as the month grid — see the CSS note.
  return `<div class="grid"
              style="grid-template-columns:3rem repeat(${days.length},minmax(0,1fr))">
      <div class="head"></div>${heads}
      <div class="daycol"><div class="allday"></div><div class="col">${hours}</div></div>
      ${days.map(dayCell).join("")}
    </div>`;
}

function render() {
  const view_el = document.getElementById("view");
  const months = ["January","February","March","April","May","June","July",
                  "August","September","October","November","December"];
  if (view === "month") {
    document.getElementById("period").textContent =
      `${months[cursor.getMonth()]} ${cursor.getFullYear()}`;
    view_el.innerHTML = renderMonth();
  } else if (view === "week") {
    const s = startOfWeek(cursor), e = addDays(s, 6);
    document.getElementById("period").textContent =
      `${months[s.getMonth()].slice(0,3)} ${s.getDate()} – ${months[e.getMonth()].slice(0,3)} ${e.getDate()}, ${e.getFullYear()}`;
    view_el.innerHTML = renderDays(Array.from({length: 7}, (_, i) => addDays(s, i)));
  } else {
    document.getElementById("period").textContent =
      `${months[cursor.getMonth()]} ${cursor.getDate()}, ${cursor.getFullYear()}`;
    view_el.innerHTML = renderDays([cursor]);
  }
  const shown = EVENTS.filter(visible).length;
  document.getElementById("count").textContent = `${shown} of ${EVENTS.length} events`;
  if (!EVENTS.length) {
    view_el.innerHTML = `<p class="empty">No events yet — upload a workbook on the
      <a href="{{ url_for('index') }}">bulk upload</a> page.</p>`;
  }
  view_el.querySelectorAll("[data-id]").forEach(n => n.onclick = () => openDetail(n.dataset.id));
  view_el.querySelectorAll(".more").forEach(n => n.onclick = () => {
    cursor = parse(n.dataset.day); view = "day"; syncViewButtons(); render();
  });
}

function openDetail(id) {
  const e = EVENTS.find(x => String(x.id) === String(id));
  if (!e) return;
  document.getElementById("dhead").style.setProperty("--c", colorOf(e.family));
  document.getElementById("dhead").innerHTML =
    `<h3>${esc(e.title)}</h3><div style="color:var(--muted);font-size:0.85rem">
       ${esc(e.venue || "Venue not resolved")}</div>`;
  const range = e.end && e.end !== e.start
    ? `${e.startLabel} – ${e.endLabel}` : e.startLabel;
  document.getElementById("dbody").innerHTML = `
    <dl>
      <dt>When</dt><dd>${range}${e.timeLabel ? " at " + e.timeLabel : ""}</dd>
      ${e.price ? `<dt>Admission</dt><dd>${esc(e.price)}</dd>` : ""}
      ${e.catNames.length ? `<dt>Categories</dt><dd>${e.catNames.map(c =>
        `<span class="tag">${esc(c)}</span>`).join("")}</dd>` : ""}
      ${e.venueId ? `<dt>Venue ID</dt><dd>${esc(e.venueId)}</dd>` : ""}
      ${e.url ? `<dt>Details</dt><dd><a href="${esc(e.url)}" target="_blank"
        rel="noopener">${esc(e.url)}</a></dd>` : ""}
      <dt>Source</dt><dd>${e.origin === "listed"
        ? "Already listed on DelawareScene" : "Found by SceneScout"}</dd>
    </dl>`;
  document.getElementById("detail").showModal();
}
document.getElementById("dclose").onclick = () => document.getElementById("detail").close();

function syncViewButtons() {
  document.querySelectorAll("#views button").forEach(b =>
    b.classList.toggle("on", b.dataset.v === view));
}
document.querySelectorAll("#views button").forEach(b => b.onclick = () => {
  view = b.dataset.v; syncViewButtons(); render();
});
document.querySelectorAll("#origins button").forEach(b => b.onclick = () => {
  originFilter = b.dataset.o;
  document.querySelectorAll("#origins button").forEach(x =>
    x.classList.toggle("on", x.dataset.o === originFilter));
  render();
});
const step = n => {
  if (view === "month") cursor = new Date(cursor.getFullYear(), cursor.getMonth() + n, 1);
  else cursor = addDays(cursor, n * (view === "week" ? 7 : 1));
  render();
};
document.getElementById("prev").onclick = () => step(-1);
document.getElementById("next").onclick = () => step(1);
document.getElementById("today").onclick = () => { cursor = new Date(); render(); };
document.getElementById("q").oninput = ev => { query = ev.target.value.trim(); render(); };
document.addEventListener("keydown", ev => {
  if (ev.target.tagName === "INPUT") return;
  if (ev.key === "ArrowLeft") step(-1);
  if (ev.key === "ArrowRight") step(1);
  if (ev.key === "t") { cursor = new Date(); render(); }
});

renderLegend();
render();
</script>"""

PAGE = """<!doctype html><meta charset="utf-8">
<title>MockScene — bulk upload</title>
<style>
  :root { --blue:#2F5384; --buff:#C9A85C; --bg:#F7F5EE; --ink:#212C39; }
  body { font-family: system-ui, sans-serif; background: var(--bg); color: var(--ink);
         margin: 0; line-height: 1.5; }
  header { background: var(--blue); color: #fff; padding: 1rem 2rem;
           display: flex; align-items: baseline; gap: 1.5rem; flex-wrap: wrap; }
  header h1 { margin: 0; font-size: 1.4rem; } header i { color: var(--buff); }
  .navlink { margin-left: auto; color: #cfe0f5; text-decoration: none; font-size: 0.9rem; }
  .navlink:hover { color: #fff; }
  main { max-width: 60rem; margin: 1.5rem auto; padding: 0 1rem; }
  .panel { background: #fff; border: 1px solid #ddd; border-radius: 8px;
           padding: 1rem 1.4rem; margin-bottom: 1.4rem; }
  .cards { display: grid; grid-template-columns: repeat(auto-fill, minmax(16rem, 1fr)); gap: 0.8rem; }
  .card { background: #fff; border: 1px solid #ddd; border-left: 4px solid var(--buff);
          border-radius: 6px; padding: 0.7rem 0.9rem; }
  .card h3 { margin: 0 0 0.3rem; font-size: 1rem; }
  .card .meta { font-size: 0.82rem; color: #555; }
  .tag { display: inline-block; background: #eef; color: var(--blue); border-radius: 99px;
         font-size: 0.7rem; padding: 0.05rem 0.5rem; margin: 0.15rem 0.15rem 0 0; }
  .err { background: #fdf0f0; border-left: 4px solid #a33; padding: 0.5rem 0.9rem;
         margin: 0.4rem 0; font-size: 0.85rem; border-radius: 4px; }
  .ok { color: #2E6B4F; font-weight: 600; }
  .dup { background: #FBF6E9; border-left: 4px solid var(--buff); padding: 0.5rem 0.9rem;
         margin: 0.5rem 0; font-size: 0.85rem; border-radius: 4px; }
  .dup summary { cursor: pointer; font-weight: 600; }
  .dup ul { margin: 0.5rem 0 0; padding-left: 1.2rem; max-height: 16rem; overflow: auto; }
  form.clear { display: inline; }
  button { background: var(--blue); color: #fff; border: 0; border-radius: 6px;
           padding: 0.45rem 1rem; cursor: pointer; }
</style>
<header><h1>Mock<span style="color:var(--buff)">Scene</span>
  <i>bulk event upload — prototype intake for DelawareScene.com</i></h1>
  <a href="{{ url_for('calendar_view') }}" class="navlink">View calendar &rarr;</a>
</header>
<main>
  <div class="panel">
    <h2>Upload spreadsheet</h2>
    <p>Same 13-column template as the real bulk-upload workbook. Rows are validated
       against the DDOA submission guidelines; continuation rows (extra performances)
       attach to the production above them.</p>
    <form method="post" action="{{ url_for('upload') }}" enctype="multipart/form-data">
      <input type="file" name="file" accept=".xlsx" required>
      <button type="submit">Upload &amp; validate</button>
    </form>
    {% if summary %}<p class="ok">{{ summary }}</p>{% endif %}
    {% for e in errors %}<div class="err">{{ e }}</div>{% endfor %}
    {% if duplicates %}
      <details class="dup"><summary>{{ dup_total }} duplicate{{ '' if dup_total == 1 else 's' }}
        skipped — already on the calendar</summary>
        <ul>{% for d in duplicates %}<li>{{ d }}</li>{% endfor %}</ul>
        {% if dup_total > duplicates|length %}
          <p><i>…and {{ dup_total - duplicates|length }} more.</i></p>{% endif %}
      </details>
    {% endif %}
  </div>
  <div class="panel">
    <h2>Calendar ({{ events|length }} listings)
      <form class="clear" method="post" action="{{ url_for('clear') }}">
        <button>clear</button></form></h2>
    <div class="cards">
    {% for ev in events %}
      <div class="card">
        <h3>{{ ev['title'] }}</h3>
        <div class="meta">
          {{ ev['start_date'] }}{% if ev['start_time'] %} @ {{ ev['start_time'] }}{% endif %}
          {% if ev['end_date'] %} — through {{ ev['end_date'] }}{% endif %}<br>
          venue #{{ ev['venue_id'] or '?' }}{% if ev['price'] %} · {{ ev['price'] }}{% endif %}
          {% if ev['url'] %} · <a href="{{ ev['url'] }}">details</a>{% endif %}
        </div>
        <div>{% for c in ev['cats'] %}<span class="tag">{{ c }}</span>{% endfor %}</div>
      </div>
    {% endfor %}
    </div>
  </div>
</main>"""


def _coerce_cell(key, value):
    """Normalize a spreadsheet cell to the string shape validate_row expects.

    Real submitters' workbooks carry Excel date/time-typed cells and lowercase
    or uppercase meridiems; those are valid submissions, so accept them rather
    than rejecting on formatting the exporter happens not to produce.
    """
    import datetime as dt

    from scenescout.export import fmt_time

    if value is None or value == "":
        return None
    if key in ("start date", "end date"):
        if isinstance(value, (dt.datetime, dt.date)):
            return value.strftime("%m/%d/%Y")
    if key == "start time":
        if isinstance(value, dt.datetime):
            return fmt_time(value.strftime("%H:%M"))
        if isinstance(value, dt.time):
            return fmt_time(f"{value.hour:02d}:{value.minute:02d}")
        text = str(value).strip()
        m = re.match(r"^(\d{1,2}):(\d{2})\s*([ap])\.?\s*m\.?$", text, re.IGNORECASE)
        if m:
            return f"{int(m[1])}:{m[2]} {m[3].lower()}.m."
        return text
    if key in ("low price", "high price"):
        text = str(value).strip()
        if text.upper() == "FREE":
            return "FREE"
        try:
            f = float(text)
        except ValueError:
            return text
        # Whole dollars only; anything with cents stays a string so
        # validate_row rejects it, matching the exporter's rule.
        return int(f) if f.is_integer() else text
    return str(value).strip()


SEED_XLSX = ROOT / "assets" / "DelawareScene Currently Listed Events.xlsx"
DEMO_XLSX = ROOT / "assets" / "scenescout-export.xlsx"
VENUES_CSV = ROOT / "assets" / "delawarescene-venues.csv"
COMBINED_XLSX = ROOT / "out" / "delawarescene-calendar.xlsx"
COMBINED_HEADERS = ["Title", "Venue", "Venue ID", "Categories", "Start Date",
                    "Start Time", "End Date", "Admission", "URL", "Source"]

COLUMNS = {
    "venue_id": "TEXT", "venue_name": "TEXT", "title": "TEXT",
    "categories": "TEXT", "url": "TEXT", "price": "TEXT",
    "start_date": "TEXT", "start_time": "TEXT", "end_date": "TEXT",
    "start_iso": "TEXT", "end_iso": "TEXT", "origin": "TEXT", "added_at": "TEXT",
}


def _db():
    """The calendar is a persistent store, not a scratch buffer.

    It is seeded once from DelawareScene's own "currently listed" export, and
    every upload merges into it, so reopening the site shows what is already
    on the calendar plus everything SceneScout has contributed so far.
    """
    DB.parent.mkdir(exist_ok=True)
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    cols = ", ".join(f"{n} {t}" for n, t in COLUMNS.items())
    conn.execute(f"CREATE TABLE IF NOT EXISTS mock_listings (id INTEGER PRIMARY KEY, {cols})")
    existing = {r["name"] for r in conn.execute("PRAGMA table_info(mock_listings)")}
    for name, decl in COLUMNS.items():
        if name not in existing:
            conn.execute(f"ALTER TABLE mock_listings ADD COLUMN {name} {decl}")
    conn.commit()
    _seed(conn)
    return conn


def _seed(conn):
    """Load DelawareScene's currently-listed events the first time only."""
    if conn.execute("SELECT COUNT(*) c FROM mock_listings").fetchone()["c"]:
        return
    if not SEED_XLSX.exists():
        return
    import datetime as dt

    wb = load_workbook(SEED_XLSX, read_only=True)
    rows = wb.active.iter_rows(values_only=True)
    next(rows)  # header: title, venue, start, end
    n, collapsed = 0, 0
    for values in rows:
        title, venue, start, end = (list(values) + [None] * 4)[:4]
        if not title:
            continue
        s_iso = start.date().isoformat() if isinstance(start, dt.datetime) else _iso_date(start)
        e_iso = end.date().isoformat() if isinstance(end, dt.datetime) else _iso_date(end)
        if not s_iso:
            continue
        title = str(title).strip()
        venue = str(venue).strip() if venue else None
        # The source export itself contains repeated rows — the same show at
        # the same venue on the same date appears twice. Collapse them here so
        # the calendar and the duplicate check both start from clean data.
        if find_duplicate(conn, title, venue, s_iso, e_iso, {}):
            collapsed += 1
            continue
        conn.execute(
            "INSERT INTO mock_listings (title, venue_name, start_date, end_date, "
            "start_iso, end_iso, origin, added_at) VALUES (?,?,?,?,?,?, 'listed', ?)",
            (title, venue, _us_date(s_iso), _us_date(e_iso), s_iso, e_iso, _now()),
        )
        n += 1
    conn.commit()
    if n:
        write_combined(conn)
        note = f" ({collapsed} duplicate rows in the source export collapsed)" if collapsed else ""
        print(f"[mock_site] seeded {n} events already listed on DelawareScene{note}")
    _preload(conn)


def _preload(conn):
    """Optionally ingest a shipped pipeline export on first boot.

    A deployed demo has nobody to upload a workbook, and a calendar showing
    only what DelawareScene already lists would hide the entire contribution.
    Off by default so a local run still starts from DelawareScene's own data
    and the upload page has something to demonstrate.
    """
    if os.environ.get("SCENESCOUT_PRELOAD", "").lower() not in ("1", "true", "yes"):
        return
    if not DEMO_XLSX.exists():
        print(f"[mock_site] SCENESCOUT_PRELOAD set but {DEMO_XLSX.name} is missing")
        return
    accepted, errors, duplicates = ingest_workbook(conn, DEMO_XLSX, _venue_names())
    print(f"[mock_site] preloaded {accepted} events SceneScout found "
          f"({len(duplicates)} already listed, {len(errors)} rejected)")


def _now():
    import datetime as dt

    return dt.datetime.now().isoformat(timespec="seconds")


def _us_date(iso):
    if not iso:
        return None
    y, m, d = iso.split("-")
    return f"{m}/{d}/{y}"


def write_combined(conn):
    """One workbook holding the whole calendar: what DelawareScene already
    lists plus everything SceneScout has added."""
    from openpyxl import Workbook

    COMBINED_XLSX.parent.mkdir(exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "calendar"
    ws.append(COMBINED_HEADERS)
    for r in conn.execute(
        "SELECT * FROM mock_listings ORDER BY start_iso, start_time, title"
    ):
        cats = ", ".join(CATEGORY_NAMES.get(int(c), c)
                         for c in (r["categories"] or "").split(",") if c.strip().isdigit())
        ws.append([
            r["title"], r["venue_name"], r["venue_id"], cats or r["categories"],
            r["start_date"], r["start_time"], r["end_date"], r["price"], r["url"],
            "Already listed" if r["origin"] == "listed" else "SceneScout",
        ])
    wb.save(COMBINED_XLSX)
    return COMBINED_XLSX


def _fetch_events(conn):
    rows = conn.execute(
        "SELECT * FROM mock_listings ORDER BY start_iso, start_time"
    ).fetchall()
    events = []
    for r in rows:
        cats = []
        for c in (r["categories"] or "").split(","):
            if c.strip().isdigit():
                cats.append(CATEGORY_NAMES.get(int(c), c))
        events.append({**dict(r), "cats": cats})
    return events


def _calendar_events(conn):
    """Shape stored listings for the calendar: ISO dates, minutes past midnight
    for positioning, and a discipline key for colour."""
    import json as _json

    venues = _venue_names()
    out = []
    for r in conn.execute("SELECT * FROM mock_listings"):
        start = _iso_date(r["start_date"])
        if not start:
            continue
        end = _iso_date(r["end_date"])
        cats = [int(c) for c in (r["categories"] or "").split(",") if c.strip().isdigit()]
        out.append({
            "id": r["id"],
            "title": r["title"],
            "venue": venues.get(str(r["venue_id"] or "").strip()),
            "venueId": r["venue_id"],
            "start": start,
            "end": end,
            "startLabel": _pretty_date(start),
            "endLabel": _pretty_date(end) if end else None,
            "timeLabel": r["start_time"],
            "minutes": _minutes(r["start_time"]),
            "price": r["price"],
            "url": r["url"],
            "cats": cats,
            "catNames": [CATEGORY_NAMES.get(c, str(c)) for c in cats],
            "family": family_of(cats),
            "origin": r["origin"] or "scenescout",
        })
    # Sort on the ISO dates: the stored MM/DD/YYYY strings sort lexically,
    # which would put January 2027 before August 2026.
    out.sort(key=lambda e: (e["start"], e["minutes"] if e["minutes"] is not None else -1))
    families = {}
    for key, label, color, _ in FAMILIES:
        count = sum(1 for e in out if e["family"] == key)
        if count:
            families[key] = {"label": label, "color": color, "count": count}
    return out, families, _json


def find_duplicate(conn, title, venue, start_iso, end_iso, venues):
    """Return the row this event duplicates, or None.

    Uses the same scorer the pipeline's dedupe stage uses, so the portal and
    the pipeline agree on what counts as the same event. Candidates are
    blocked on date overlap first, which keeps this cheap.
    """
    if not start_iso:
        return None
    # Guard against an inverted range, which would make the overlap test
    # match nothing — including the event itself.
    end_iso = max(end_iso or start_iso, start_iso)
    rows = conn.execute(
        "SELECT * FROM mock_listings WHERE start_iso <= ? AND "
        "COALESCE(end_iso, start_iso) >= ?", (end_iso, start_iso)
    ).fetchall()
    best, best_score, used_venue = None, 0.0, False
    for r in rows:
        other_venue = r["venue_name"] or venues.get(str(r["venue_id"] or "").strip())
        score, uv = score_pair(title, venue, r["title"], other_venue)
        if score > best_score:
            best, best_score, used_venue = r, score, uv
    bar = DUPE_SCORE if used_venue else TITLE_ONLY_DUPE
    return best if best_score >= bar else None


def _venue_names():
    """VenueID -> organization name, from the DelawareScene directory the
    pipeline scraped. The workbook carries only IDs, as the real importer
    expects, so the calendar resolves them for display.

    Falls back to the directory snapshot committed alongside the site, so a
    fresh clone or a deployed instance still shows venue names without having
    run the pipeline first.
    """
    scene_db = ROOT / "data" / "scenescout.db"
    if scene_db.exists():
        try:
            c = sqlite3.connect(f"file:{scene_db}?mode=ro", uri=True)
            rows = c.execute("SELECT scene_id, name FROM scene_orgs").fetchall()
            c.close()
            if rows:
                return {str(sid): name for sid, name in rows}
        except sqlite3.Error:
            pass
    if not VENUES_CSV.exists():
        return {}
    import csv

    with VENUES_CSV.open(encoding="utf-8", newline="") as fh:
        reader = csv.DictReader(fh)
        return {r["venue_id"]: r["name"] for r in reader if r.get("venue_id")}


def _iso_date(value):
    if not value:
        return None
    m = re.match(r"^(\d{2})/(\d{2})/(\d{4})$", str(value).strip())
    return f"{m[3]}-{m[1]}-{m[2]}" if m else None


def _pretty_date(iso):
    import datetime as dt

    if not iso:
        return None
    d = dt.date.fromisoformat(iso)
    return d.strftime("%a %b %-d, %Y") if sys.platform != "win32" else d.strftime("%a %b %d, %Y")


def _minutes(time_label):
    """'7:30 p.m.' -> 1170. None for all-day events."""
    if not time_label:
        return None
    m = re.match(r"^(\d{1,2}):(\d{2})\s*([ap])\.?m\.?$", str(time_label).strip(), re.I)
    if not m:
        return None
    hour = int(m[1]) % 12 + (12 if m[3].lower() == "p" else 0)
    return hour * 60 + int(m[2])


@app.get("/calendar")
def calendar_view():
    conn = _db()
    events, families, _json = _calendar_events(conn)
    return render_template_string(
        CALENDAR_PAGE,
        events_json=_json.dumps(events),
        families_json=_json.dumps(families),
        other_color=OTHER_COLOR,
    )


@app.get("/")
def index():
    conn = _db()
    total = conn.execute("SELECT COUNT(*) c FROM mock_listings").fetchone()["c"]
    listed = conn.execute("SELECT COUNT(*) c FROM mock_listings WHERE origin='listed'"
                          ).fetchone()["c"]
    summary = (f"Calendar holds {total} events — {listed} already listed on "
               f"DelawareScene, {total - listed} added by SceneScout.") if total else None
    return render_template_string(PAGE, events=_fetch_events(conn), errors=[],
                                  summary=summary, duplicates=[], dup_total=0)


def ingest_workbook(conn, source, venues):
    """Merge a 13-column bulk-upload workbook into the calendar.

    Shared by the upload route and the boot-time preload, so a deployed demo
    ingests results through exactly the code path a real upload takes.
    Returns (accepted, errors, duplicates).
    """
    errors, accepted, duplicates = [], 0, []
    if not source:
        errors.append("no file")
    else:
        try:
            wb = load_workbook(source, read_only=True)
            ws = wb.active
            rows = ws.iter_rows(values_only=True)
            header = [str(h).strip() if h else "" for h in next(rows)]
            if [h.lower() for h in header[:13]] != [h.lower() for h in HEADERS]:
                errors.append(f"header mismatch — expected the 13-column template, got {header[:13]}")
            else:
                last_full = None
                for i, values in enumerate(rows, start=2):
                    row = dict(zip(HEADERS, list(values)[:13]))
                    if not any(v not in (None, "") for v in row.values()):
                        continue
                    is_continuation = not row["title of program"] and row["start date"] and last_full
                    if is_continuation:
                        merged = dict(last_full)
                        merged["start date"] = row["start date"]
                        merged["start time"] = row["start time"]
                        merged["ticket URL"] = row["ticket URL"] or merged["ticket URL"]
                        # A continuation row is one performance, not another
                        # copy of the run: inheriting the parent's end date
                        # would give a later performance an end date before
                        # its own start.
                        merged["end date"] = None
                        row = merged
                    row = {k: _coerce_cell(k, v) for k, v in row.items()}
                    problems = validate_row(row)
                    if problems:
                        errors.append(f"row {i}: {'; '.join(problems)}")
                        # Drop the parent so its continuation rows become
                        # orphans and are reported, rather than silently
                        # attaching to the previous production.
                        if not is_continuation:
                            last_full = None
                        continue
                    if not is_continuation:
                        last_full = row
                    price = None
                    if row["low price"] == "FREE":
                        price = "FREE"
                    elif row["low price"] is not None:
                        price = f"${row['low price']}"
                        if row["high price"]:
                            price += f"–${row['high price']}"
                    start_iso = _iso_date(row["start date"])
                    end_iso = _iso_date(row["end date"])
                    venue_name = venues.get(str(row["venue ID"] or "").strip())

                    # Check against the whole calendar — DelawareScene's own
                    # listings and everything uploaded before — so re-uploading
                    # the same workbook adds nothing.
                    dup = find_duplicate(conn, row["title of program"], venue_name,
                                         start_iso, end_iso, venues)
                    if dup:
                        where = "already on DelawareScene" if dup["origin"] == "listed" \
                            else "already imported"
                        duplicates.append(
                            f"{row['title of program']} ({row['start date']}) — {where}")
                        continue

                    conn.execute(
                        "INSERT INTO mock_listings (venue_id, venue_name, title, categories, "
                        "url, price, start_date, start_time, end_date, start_iso, end_iso, "
                        "origin, added_at) VALUES (?,?,?,?,?,?,?,?,?,?,?, 'scenescout', ?)",
                        (row["venue ID"], venue_name, row["title of program"],
                         row["categories"], row["URL"], price, row["start date"],
                         row["start time"], row["end date"], start_iso, end_iso, _now()),
                    )
                    accepted += 1
                conn.commit()
                write_combined(conn)
        except Exception as e:  # noqa: BLE001
            errors.append(f"could not read workbook: {e}")
    return accepted, errors, duplicates


@app.post("/upload")
def upload():
    conn = _db()
    accepted, errors, duplicates = ingest_workbook(
        conn, request.files.get("file"), _venue_names()
    )
    total = conn.execute("SELECT COUNT(*) c FROM mock_listings").fetchone()["c"]
    summary = None
    if accepted or errors or duplicates:
        summary = (f"{accepted} new events added · {len(duplicates)} skipped as duplicates · "
                   f"{len(errors)} rejected. Calendar now holds {total} events; "
                   f"saved to out/{COMBINED_XLSX.name}.")
    return render_template_string(PAGE, events=_fetch_events(conn), errors=errors,
                                  summary=summary, duplicates=duplicates[:40],
                                  dup_total=len(duplicates))


@app.post("/clear")
def clear():
    """Wipe everything and re-seed from DelawareScene's currently-listed export."""
    conn = _db()
    conn.execute("DELETE FROM mock_listings")
    conn.commit()
    _seed(conn)
    write_combined(conn)
    return redirect(url_for("index"))


if __name__ == "__main__":
    # 127.0.0.1 locally; a container/instance must bind 0.0.0.0 to be reachable.
    app.run(host=os.environ.get("HOST", "127.0.0.1"),
            port=int(os.environ.get("PORT", "5050")), debug=False)
